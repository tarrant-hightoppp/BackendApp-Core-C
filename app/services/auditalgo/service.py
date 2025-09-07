from __future__ import annotations
from typing import List, Optional, Dict
from datetime import datetime
import io
import re

from sqlalchemy import select, func, asc, or_
from sqlalchemy.orm import Session

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# адаптирай пътя при нужда
from app.models.operation import AccountingOperation

# S3 адаптер към наличния services.s3
try:
    from app.services import s3 as s3_module
except Exception:  # pragma: no cover
    s3_module = None

class S3Adapter:
    def __init__(self):
        if s3_module is None:
            raise RuntimeError("S3 module not available")

    def upload_bytes(self, key: str, data: bytes, content_type: str = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet") -> str:
        if hasattr(s3_module, "upload_file"):
            s3_module.upload_file(key, data)
            return key
        raise RuntimeError("s3.upload_file(key, bytes) not found; add adapter as needed")

    def presign(self, key: str, expiration: int = 3600) -> Optional[str]:
        if hasattr(s3_module, "generate_presigned_url"):
            return s3_module.generate_presigned_url(key, expiration)
        return None

# колони за експорт (съобразени с модела)
EXPORT_COLUMNS = [
    "operation_date",
    "document_type",
    "document_number",
    "debit_account",
    "credit_account",
    "amount",
    "description",
    "partner_name",
    "analytical_debit",
    "analytical_credit",
    "account_name",
    "file_id",
    "template_type",
    "created_at",
    "id",
]


def canonicalize_account(raw: Optional[str]) -> str:
    """Нормализира сметки: '121 2'/'121-2'/'121.2' → '121/2', '401.4' → '401/4'."""
    if not raw:
        return ""
    s = str(raw).strip()
    s = re.sub(r"[\s\-.]+", "/", s)
    s = re.sub(r"/+", "/", s)
    return s


def _row_to_list(r: AccountingOperation) -> List:
    return [
        getattr(r, "operation_date", None),
        getattr(r, "document_type", None),
        getattr(r, "document_number", None),
        getattr(r, "debit_account", None),
        getattr(r, "credit_account", None),
        float(getattr(r, "amount", 0) or 0),
        getattr(r, "description", None),
        getattr(r, "partner_name", None),
        getattr(r, "analytical_debit", None),
        getattr(r, "analytical_credit", None),
        getattr(r, "account_name", None),
        getattr(r, "file_id", None),
        getattr(r, "template_type", None),
        getattr(r, "created_at", None),
        getattr(r, "id", None),
    ]


def _export_xlsx(rows: List[AccountingOperation]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "operations"
    # header
    ws.append(EXPORT_COLUMNS)
    # data
    for r in rows:
        ws.append(_row_to_list(r))
    # автом. ширина
    for col_idx, col_name in enumerate(EXPORT_COLUMNS, start=1):
        max_len = len(str(col_name))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            v = row[0].value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(60, max_len + 2))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _get_col(side: str):
    if side == "debit":
        return getattr(AccountingOperation, "debit_account")
    if side == "credit":
        return getattr(AccountingOperation, "credit_account")
    raise ValueError("side must be 'debit' or 'credit'")


def _fetch_distinct_accounts(db: Session, side: str) -> List[str]:
    col = _get_col(side)
    stmt = select(func.distinct(col)).where(col.isnot(None))
    rows = db.execute(stmt).scalars().all()
    seen = {}
    for r in rows:
        if not r:
            continue
        seen[canonicalize_account(str(r))] = True
    return sorted(seen.keys())


def _fetch_ops_for_account(db: Session, side: str, account: str, prefix_mode: bool) -> List[AccountingOperation]:
    col = _get_col(side)
    canon = canonicalize_account(account)
    if prefix_mode:
        patterns = [
            f"{canon}%",
            canon.replace("/", " ") + "%",
            canon.replace("/", "-") + "%",
            canon.replace("/", ".") + "%",
        ]
        stmt = select(AccountingOperation).where(or_(*[col.like(p) for p in patterns]))
    else:
        variants = {canon, canon.replace("/", " "), canon.replace("/", "-"), canon.replace("/", ".")}
        stmt = select(AccountingOperation).where(col.in_(list(variants)))
    stmt = stmt.order_by(asc(AccountingOperation.operation_date), asc(AccountingOperation.id))
    return db.execute(stmt).scalars().all()


def _select_top_coverage(rows: List[AccountingOperation], coverage: float = 0.80, by_abs: bool = False) -> List[AccountingOperation]:
    if not rows:
        return []
    if by_abs:
        total = sum(abs(r.amount or 0) for r in rows)
        ordered = sorted(rows, key=lambda r: abs(r.amount or 0), reverse=True)
        target = coverage * total
        acc = 0.0
        take = []
        for r in ordered:
            acc += abs(r.amount or 0)
            take.append(r)
            if acc >= target:
                break
        return take
    else:
        total = sum((r.amount or 0) for r in rows)
        ordered = sorted(rows, key=lambda r: (r.amount or 0), reverse=True)
        target = coverage * total
        acc = 0.0
        take = []
        for r in ordered:
            acc += (r.amount or 0)
            take.append(r)
            if acc >= target:
                break
        return take


def _process_side(
    db: Session,
    s3: S3Adapter,
    side: str,
    *,
    coverage: float,
    coverage_by_abs: bool,
    max_full: int,
    prefix_mode: bool,
    s3_prefix: str,
    make_presigned_urls: bool,
) -> List[Dict]:
    results: List[Dict] = []
    accounts = _fetch_distinct_accounts(db, side)
    now = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")

    for acc in accounts:
        rows = _fetch_ops_for_account(db, side, acc, prefix_mode=prefix_mode)
        n = len(rows)
        if n == 0:
            continue

        total_amount = float(sum((r.amount or 0) for r in rows)) if not coverage_by_abs else float(sum(abs(r.amount or 0) for r in rows))

        if n <= max_full:
            export_rows = rows
            export_type = "full"
        else:
            export_rows = _select_top_coverage(rows, coverage=coverage, by_abs=coverage_by_abs)
            export_type = "top80"

        xlsx_bytes = _export_xlsx(export_rows)
        key = f"{s3_prefix}/{now}/{side}_{acc}_{export_type}.xlsx"
        s3_key = s3.upload_bytes(key, xlsx_bytes)
        s3_url = s3.presign(s3_key) if make_presigned_urls else None

        results.append({
            "side": side,
            "account": acc,
            "canonical_account": canon,
            "match_mode": "prefix" if prefix_mode else "exact",
            "total_rows": n,
            "exported_rows": len(export_rows),
            "total_amount": total_amount,
            "coverage_target": coverage,
            "coverage_by_abs": coverage_by_abs,
            "s3_key": s3_key,
            "s3_url": s3_url,
            "export_type": export_type,
        })

    return results


def run_audit(
    db: Session,
    s3: Optional[S3Adapter] = None,
    *,
    coverage: float = 0.80,
    coverage_by_abs: bool = False,
    max_full: int = 30,
    prefix_mode: bool = False,
    process_sides: Optional[List[str]] = None,  # по подразбиране ["debit","credit"]
    s3_prefix: str = "audit",
    make_presigned_urls: bool = True,
) -> List[Dict]:
    """Обхожда всички уникални сметки по дебит, после по кредит. Генерира XLSX и качва в S3."""
    if s3 is None:
        s3 = S3Adapter()
    if not process_sides:
        process_sides = ["debit", "credit"]

    all_results: List[Dict] = []
    for side in process_sides:
        all_results.extend(
            _process_side(
                db,
                s3,
                side,
                coverage=coverage,
                coverage_by_abs=coverage_by_abs,
                max_full=max_full,
                prefix_mode=prefix_mode,
                s3_prefix=s3_prefix,
                make_presigned_urls=make_presigned_urls,
            )
        )
    return all_results