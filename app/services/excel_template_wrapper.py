import io
import pandas as pd
from typing import Dict, Any, List, Optional, BinaryIO, Union
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import WHITE
from datetime import datetime

from app.core.config import settings
from app.services.s3 import S3Service


class ExcelTemplateWrapper:
    """Service for wrapping exported Excel files in a predefined template"""
    
    def __init__(self):
        """Initialize the template wrapper service"""
        self.s3_service = S3Service()
        
    def _create_template_workbook(self,
                                 company_name: str = "Форт България ЕООД",
                                 year: str = None,
                                 auditor_name: str = "ПРИМА ФИНАНС КОНСУЛТИНГ ЕООД",
                                 audit_approach: str = "statistical") -> Workbook:
        """
        Create a new workbook with the C700 template structure
        
        Args:
            company_name: Name of the company to include in the template
            year: Year to include in the template (default: current year)
            auditor_name: Name of the auditor company
            
        Returns:
            Workbook object with the template structure
        """
        # If year is not provided, use current year
        if year is None:
            year = str(datetime.now().year)
            
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Define styles
        header_font = Font(name='Calibri', size=11, bold=True)
        normal_font = Font(name='Calibri', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        # Define a thicker border for the C-D column separation
        thick_right_border = Border(
            left=Side(style='thin'),
            right=Side(style='thick'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        # Colors based on the C700.xlsx template - exact color codes
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        teal_fill = PatternFill(start_color="33CCCC", end_color="33CCCC", fill_type="solid")  # Exact teal color for header rows
        light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        footer_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White for footer
        
        # Set column widths - match the exact widths from the template
        column_widths = {
            'A': 15.43,   # № по ред
            'B': 22.14,  # Документ №/Дата
            'C': 10.0,   # Рег. №
            'D': 10.0,   # Дт с/ка
            'E': 25.71,  # Аналитична сметка (Дт)
            'F': 10.0,   # Кт с/ка
            'G': 25.71,  # Аналитична сметка (Кт)
            'H': 15.0,   # Сума
            'I': 30.0,   # Обяснение/Обоснование
            'J': 15.0,   # Установена сума при одита
            'K': 15.0,   # Отклонение
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
            
        # Set row heights - match the exact heights from the template
        row_heights = {
            # Header section
            1: 38, 2: 20.25, 3: 50.0, 4: 40, 5: 40, 6: 40, 7: 30.0, 8: 20.25,
            # Audit purpose section
            9: 20.25, 10: 30.0, 11: 20.25, 12: 20.25, 13: 30.0,
            # Approach section
            15: 20.25, 16: 20.25, 17: 20.25, 18: 20.25,
            # Operations header
            26: 30.0,
            #Footer \\^.^//
            72: 30, 74: 30, 79: 30,84: 35,
        }
        
        # Set specific row heights
        for row, height in row_heights.items():
            ws.row_dimensions[row].height = height
            
        # Set default height for other rows
        for i in range(1, 100):
            if i not in row_heights:
                if i >= 71:  # Footer section
                    ws.row_dimensions[i].height = 18.0  # Taller rows for footer
                else:
                    ws.row_dimensions[i].height = 15.0
        
        # Create merged cells for the header section based on the exact structure in C700.xlsx
        # First row
        ws.merge_cells('A1:A1')
        ws.merge_cells('B1:C1')
        ws.merge_cells('D1:E1')
        ws.merge_cells('F1:G1')
        
        # Second row
        ws.merge_cells('A2:A2')
        ws.merge_cells('B2:C2')
        ws.merge_cells('D2:E2')
        ws.merge_cells('F2:G2')
        
        # Third row
        ws.merge_cells('A3:A3')
        ws.merge_cells('B3:C3')
        ws.merge_cells('D3:E3')
        ws.merge_cells('F3:G3')
        
        # Fourth row
        ws.merge_cells('A4:A4')
        ws.merge_cells('B4:C4')
        ws.merge_cells('D4:E4')
        ws.merge_cells('F4:G4')
        
        # Fifth row
        ws.merge_cells('A5:A5')
        ws.merge_cells('B5:C5')
        ws.merge_cells('D5:E5')
        ws.merge_cells('F5:G5')
        
        # Sixth row
        ws.merge_cells('A6:A6')
        ws.merge_cells('B6:C6')
        ws.merge_cells('D6:E6')
        ws.merge_cells('F6:G6')
        
        # Seventh row
        ws.merge_cells('A7:C7')
        ws.merge_cells('D7:G7')
        
        # Eighth row
        ws.merge_cells('A8:C8')
        ws.merge_cells('D8:G8')
        
        # Header section - First part (rows 1-8)
        # Row 1
        ws['A1'] = "ОДИТЪТ СЕ ИЗВЪРШВА ОТ"
        ws['B1'] = auditor_name
        ws['D1'] = "ДОКУМЕНТ"
        ws['F1'] = "С 700  Тест  по  същество  на   приходи"
        
        # Row 2
        ws['A2'] = "АДРЕС"
        ws['B2'] = "СОФИЯ, УЛ. ЦАР ШИШМАН 17"
        ws['D2'] = "ОДИТИРАНО ПРЕДПРИЯТИЕ"
        ws['F2'] = company_name
        
        # Row 3
        ws['A3'] = "БУЛСТАТ"
        ws['B3'] = "121100122"
        ws['D3'] = "АДРЕС"
        ws['F3'] = "гр.София, жк Младост, Бизнес\nЦентър \"Капитал Форт\", бул. \"Цариградско шосе\" No 90, ет. 13"
        
        # Row 4
        ws['A4'] = "РЪКОВОДИТЕЛ ЕКИП"
        ws['B4'] = "ВАСИЛ КРЪСТЕВ КАЛАЙДЖИЕВ"
        ws['D4'] = "БУЛСТАТ"
        ws['F4'] = "203576042"
        
        # Row 5
        ws['A5'] = "ДИПЛОМА НОМЕР"
        ws['B5'] = "409"
        ws['D5'] = "ПРОВЕРЯВАН ПЕРИОД"
        ws['F5'] = year
        
        # Row 6
        ws['A6'] = "ДАТА НА ИЗГОТВЯНЕ"
        ws['B6'] = datetime.now().strftime("%d/%m/%Y")
        ws['D6'] = "ДАТА НА ПРОВЕРКА"
        ws['F6'] = datetime.now().strftime("%d/%m/%Y")
        
        # Row 7
        ws['A7'] = "ИЗГОТВИЛ ИЛИ НАНЕСЪЛ ПОПРАВКИ (ОДИТОР/ПОМОЩНИК ОДИТОР,АСИСТЕНТ)"
        ws['D7'] = "ПРОВЕРИЛ (ОТГОВОРЕН ОДИТОР)"
        
        # Row 8
        ws['A8'] = "ИС"
        ws['D8'] = "ВК"
        
        # Apply styles to header section - teal background for header rows
        for row in range(1, 9):
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.font = normal_font
                cell.alignment = Alignment(wrap_text=True, vertical='center')
                # Apply thicker border between columns C and D in the header
                if col == 3:  # Column C
                    cell.border = thick_right_border
                else:
                    cell.border = border
                cell.fill = teal_fill  # Apply teal background to header rows
        
        # Audit purpose section - merge cells for all columns A-G
        for row in range(9, 14):
            ws.merge_cells(f'A{row}:G{row}')
        
        # Audit purpose section - content
        ws['A9'] = "Цел   на одиторската  процедура "
        ws['A10'] = "Целта  на   одиторската  процедура  е да  установи   СНОН  при  признаването ,  оценката , последваща  оценка ,  класификация  и представяне  "
        ws['A11'] = "на приходи  , включително  и  финансови  ."
        ws['A12'] = "Приложени одиторсик процедури :"
        ws['A13'] = "факт.проверка  на договори и др.документация;повторно изчисление на салдо ;равнение на сметката с ФО ;проучващи запитвания"
        
        # Apply styles to audit purpose section
        for row in range(9, 14):
            cell = ws.cell(row=row, column=1)
            cell.font = normal_font
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')  # Center horizontally
            cell.border = border
        
        # Approach section - merge cells
        # Row 15 - header
        ws.merge_cells('C15:G15')
        
        # Rows 16-18 - options
        for row in range(16, 19):
            ws.merge_cells(f'C{row}:G{row}')
        
        # Approach section - content
        ws['C15'] = "  Избран подход за тест по  същество "
        ws['C16'] = "проверка на 100 %  на  популация "
        ws['C17'] = "проверка   на  избрани  обекти  на  популация  "
        # Set the X mark based on the audit approach
        if audit_approach == "full":
            ws['A16'] = "X"  # 100% population check
            ws['A17'] = ""
            ws['A18'] = ""
        elif audit_approach == "selected":
            ws['A16'] = ""
            ws['A17'] = "X"  # Check of selected population objects
            ws['A18'] = ""
        else:  # statistical (default)
            ws['A16'] = ""
            ws['A17'] = ""
            ws['A18'] = "X"  # Statistical audit sampling
        ws['C18'] = "одиторска  извадка   - статистическа "
        
        # Apply styles to approach section
        # Style for the header
        cell = ws.cell(row=15, column=3)
        cell.font = normal_font
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.border = border
        
        # Style for the checkboxes
        for row in [16, 17, 18]:
            cell = ws.cell(row=row, column=1)
            cell.font = normal_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Style for the options
        for row in range(16, 19):
            cell = ws.cell(row=row, column=3)
            cell.font = normal_font
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.border = border
        
        # Skip rows 19-25 (empty in the template)
        
        # Skip rows 19-25 (empty in the template)
        
        # Table header for operations - row 26
        headers = [
            "Документ №/Дата", "Рег. №", "Дт с/ка", "Аналитична сметка",
            "Кт с/ка", "Аналитична сметка", "Сума", "Обяснителен текст",
            "Установена сума   при одита ", "Отклонение  "
        ]
        
        # Add column headers for operations table (row 26)
        for i, header in enumerate(["№ по ред"] + headers):
            cell = ws.cell(row=26, column=i+1)
            cell.value = header
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            cell.fill = teal_fill  # Use teal color for the operations header row
            
        # Add empty rows for operations
        # For "full" audit approach, we need more rows - allocate 500 rows to be safe
        max_row = 500 if audit_approach == "full" else 71
        for row in range(27, max_row):
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.border = border
        
        # Add conclusion section headers
        # For "full" audit approach, the conclusion section starts after the operations
        conclusion_start_row = max_row
        
        # Create the conclusion table header
        ws.merge_cells(f'A{conclusion_start_row}:K{conclusion_start_row}')
        
        # Merge A-B for the label columns
        for i in range(1, 15):
            row = conclusion_start_row + i
            ws.merge_cells(f'A{row}:B{row}')
        
        # Don't merge C-G for content cells - we need separate columns for each account
        # Instead, we'll create a table-like structure with individual cells
        
        # Conclusion section content - headers only (values will be populated dynamically)
        ws[f'A{conclusion_start_row}'] = "ЗАКЛЮЧЕНИЯ :"
        ws[f'A{conclusion_start_row+1}'] = "Обща  сума/брой   на  проверени   документи "
        ws[f'A{conclusion_start_row+2}'] = "Обща  сума / брой на обекти в  популация "
        ws[f'A{conclusion_start_row+3}'] = "Равнение на  ст/ст на популация с оборотна ведомости; глава  книга ; ФО "
        ws[f'A{conclusion_start_row+4}'] = "Проектиране на грешката "
        ws[f'A{conclusion_start_row+5}'] = "Проектиране на грешката "
        ws[f'C{conclusion_start_row+5}'] = "НЕПРИЛОЖИМО"
        ws[f'A{conclusion_start_row+6}'] = ""
        ws[f'A{conclusion_start_row+7}'] = ""
        ws[f'A{conclusion_start_row+8}'] = "Констатирани  СНОН "
        ws[f'C{conclusion_start_row+8}'] = "Не са констатирани съществени неточности, отклонения и несъответствия при осчетоводяване на продажбите."
        ws[f'A{conclusion_start_row+13}'] = "Други  заключения "
        # Will be populated dynamically
        
        # Apply styles to conclusion section - with proper formatting
        # Header row (ЗАКЛЮЧЕНИЯ) with improved formatting
        for col in range(1, 12):
            cell = ws.cell(row=conclusion_start_row, column=col)
            cell.font = Font(name='Calibri', size=12, bold=True)  # Slightly larger font
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')  # Center align
            cell.border = border
            cell.fill = teal_fill  # Use teal color for the conclusion header
        
        # Content rows with improved formatting
        for i in range(1, 20):  # Format 20 rows after conclusion header
            row = conclusion_start_row + i
            
            # Format label column (A-B merged)
            label_cell = ws.cell(row=row, column=1)
            if not label_cell.font:
                label_cell.font = normal_font
            label_cell.alignment = Alignment(wrap_text=True, vertical='center')
            label_cell.border = border
            
            # Apply different formatting to different sections
            if i in [1, 2, 3]:  # First section - Account totals
                label_cell.fill = light_gray_fill
                label_cell.font = Font(name='Calibri', size=11, bold=True)  # Make labels bold
            elif i in [4, 5, 6, 7]:  # Проектиране на грешката section
                label_cell.fill = light_gray_fill
                if i == 4:  # Section header
                    label_cell.font = Font(name='Calibri', size=11, bold=True)
            elif i in [8, 9, 10, 11, 12]:  # Констатирани СНОН section
                label_cell.fill = light_gray_fill
                if i == 8:  # Section header
                    label_cell.font = Font(name='Calibri', size=11, bold=True)
            elif i == 13:  # Други заключения section
                label_cell.fill = light_gray_fill
                label_cell.font = Font(name='Calibri', size=11, bold=True)
            
            # Format content cells (C-K) - individual cells for each account
            for col in range(3, 12):
                cell = ws.cell(row=row, column=col)
                if not cell.font:
                    cell.font = normal_font
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                cell.border = border
                cell.fill = footer_fill
        
        return wb
        
    def wrap_excel_with_template(self,
                                excel_content: Union[BinaryIO, bytes],
                                company_name: str = "Форт България ЕООД",
                                year: str = None,
                                audit_approach: str = "statistical",
                                account_type: str = None) -> io.BytesIO:
        """
        Wrap an Excel file with operations data in a predefined template
        
        Args:
            excel_content: Content of the Excel file to wrap (file-like object or bytes)
            company_name: Name of the company to include in the template
            year: Year to include in the template (default: current year)
            audit_approach: The audit approach to use (default: "statistical")
            account_type: The type of account being analyzed ("debit" or "credit")
            
        Returns:
            BytesIO object containing the wrapped Excel file
        """
        # If excel_content is bytes, convert to file-like object
        if isinstance(excel_content, bytes):
            excel_content = io.BytesIO(excel_content)
            
        # If year is not provided, use current year
        if year is None:
            year = str(datetime.now().year)
            
        # Create the template workbook
        template_wb = self._create_template_workbook(
            company_name=company_name,
            year=year,
            audit_approach=audit_approach
        )
        
        # Load the operations data
        operations_df = pd.read_excel(excel_content)
        
        # Get the first sheet of the template
        template_sheet = template_wb.active
        
        # Determine the verification period based on operation dates
        operation_dates = []
        if "Дата" in operations_df.columns:
            for date in operations_df["Дата"]:
                if isinstance(date, datetime):
                    operation_dates.append(date)
                elif isinstance(date, str):
                    try:
                        # Try to parse the date string
                        parsed_date = datetime.strptime(date, "%d.%m.%Y")
                        operation_dates.append(parsed_date)
                    except ValueError:
                        try:
                            # Try alternative format
                            parsed_date = datetime.strptime(date, "%Y-%m-%d")
                            operation_dates.append(parsed_date)
                        except ValueError:
                            # Skip invalid dates
                            pass
        
        # Set the verification period in the template
        if operation_dates:
            # Sort dates to find min and max
            operation_dates.sort()
            start_date = operation_dates[0]
            end_date = operation_dates[-1]
            
            # Format the verification period
            verification_period = f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
            
            # Update the verification period in the template (cell F5)
            template_sheet["F5"] = verification_period
        
        # Starting row for operations data
        start_row = 27
        
        # Insert operations data into the template
        row_count = 0
        total_by_account = {}
        current_account_rows = []
        
        # For "full" audit approach, we need to include all operations
        # For other approaches, we'll limit the number of rows
        max_rows = float('inf') if audit_approach == "full" else 43  # No limit for full audit approach
        
        for i, row_data in operations_df.iterrows():
            # Skip if we've reached the maximum number of rows we can fit
            # But for "full" audit approach, we'll include all operations
            if row_count >= max_rows and audit_approach != "full":
                break
                
            row_num = start_row + row_count
            
            # Map DataFrame columns to template columns - exactly matching the original template format
            # Shift data one column to the right as requested
            
            # Column 1: № по ред (sequence number)
            if "№ по ред" in row_data:
                template_sheet.cell(row=row_num, column=1).value = row_data["№ по ред"]
            
            # Column 2: Документ №/Дата
            doc_num = row_data.get("Документ №", "")
            doc_date = row_data.get("Дата", "")
            if doc_num and doc_date:
                if isinstance(doc_date, datetime):
                    doc_date = doc_date.strftime("%d.%m.%Y")
                template_sheet.cell(row=row_num, column=2).value = f"{doc_num}, {doc_date}"
            
            # Column 3: Рег. №
            # This is usually empty in the template, but we'll map it if available
            if "Рег. №" in row_data:
                template_sheet.cell(row=row_num, column=3).value = row_data["Рег. №"]
            
            # Column 4: Дт с/ка
            debit_account = None
            if "Дт с/ка" in row_data:
                debit_account = row_data["Дт с/ка"]
                template_sheet.cell(row=row_num, column=4).value = debit_account
            
            # Column 5: Аналитична сметка (Дт)
            if "Аналитична сметка/Партньор (Дт)" in row_data:
                template_sheet.cell(row=row_num, column=5).value = row_data["Аналитична сметка/Партньор (Дт)"]
            
            # Column 6: Кт с/ка
            credit_account = None
            if "Кт с/ка" in row_data:
                credit_account = row_data["Кт с/ка"]
                template_sheet.cell(row=row_num, column=6).value = credit_account
                
                # Track totals by credit account for summary
                if credit_account not in total_by_account:
                    total_by_account[credit_account] = 0
                
                # Track rows for this account for subtotals
                current_account_rows.append((row_num, credit_account))
            
            # Column 7: Аналитична сметка (Кт)
            if "Аналитична сметка/Партньор (Кт)" in row_data:
                template_sheet.cell(row=row_num, column=7).value = row_data["Аналитична сметка/Партньор (Кт)"]
            
            # Column 8: Сума - with improved formatting
            amount = 0
            if "Сума" in row_data:
                amount = row_data["Сума"]
                amount_cell = template_sheet.cell(row=row_num, column=8)
                amount_cell.value = amount
                amount_cell.alignment = Alignment(horizontal='right', vertical='center')
                amount_cell.number_format = '#,##0.00'
                
                # Add to account total
                if credit_account in total_by_account:
                    if isinstance(amount, (int, float)):
                        total_by_account[credit_account] += amount
                    else:
                        try:
                            total_by_account[credit_account] += float(amount)
                        except (ValueError, TypeError):
                            pass  # Skip if we can't convert to float
            
            # Column 9: Обяснителен текст - with improved formatting
            if "Обяснение/Обоснование" in row_data:
                text_cell = template_sheet.cell(row=row_num, column=9)
                text_cell.value = row_data["Обяснение/Обоснование"]
                text_cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            # Column 10: Установена сума при одита - with improved formatting
            verified_cell = template_sheet.cell(row=row_num, column=10)
            if "Установена сума при одита" in row_data:
                verified_cell.value = row_data["Установена сума при одита"]
            else:
                # If not provided, use the same as the amount
                verified_cell.value = amount
            verified_cell.alignment = Alignment(horizontal='right', vertical='center')
            verified_cell.number_format = '#,##0.00'
            
            # Column 11: Отклонение - with improved formatting
            deviation_cell = template_sheet.cell(row=row_num, column=11)
            deviation_cell.value = 0.0
            deviation_cell.alignment = Alignment(horizontal='right', vertical='center')
            deviation_cell.number_format = '#,##0.00'
            
            row_count += 1
            
            # Add a subtotal row after a group of operations with the same account
            if i < len(operations_df) - 1:
                next_row = operations_df.iloc[i + 1]
                current_account = row_data.get("Кт с/ка", "")
                next_account = next_row.get("Кт с/ка", "")
                
                if current_account != next_account and current_account:
                    # This is the last row of a group, add a subtotal with improved formatting
                    subtotal_row = start_row + row_count
                    
                    # Add "Общо" label
                    total_cell = template_sheet.cell(row=subtotal_row, column=1)
                    total_cell.value = "Общо"
                    total_cell.font = Font(name='Calibri', size=11, bold=True)
                    total_cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    # Add account number with bold formatting
                    account_cell = template_sheet.cell(row=subtotal_row, column=6)
                    account_cell.value = current_account
                    account_cell.font = Font(name='Calibri', size=11, bold=True)
                    account_cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Add total amount with bold formatting and proper number format
                    if current_account in total_by_account:
                        # Format amount cell
                        amount_cell = template_sheet.cell(row=subtotal_row, column=8)
                        amount_cell.value = total_by_account[current_account]
                        amount_cell.font = Font(name='Calibri', size=11, bold=True)
                        amount_cell.alignment = Alignment(horizontal='right', vertical='center')
                        amount_cell.number_format = '#,##0.00'
                        
                        # Format verified amount cell (same as total from database)
                        verified_cell = template_sheet.cell(row=subtotal_row, column=10)
                        verified_cell.value = total_by_account[current_account]
                        verified_cell.font = Font(name='Calibri', size=11, bold=True)
                        verified_cell.alignment = Alignment(horizontal='right', vertical='center')
                        verified_cell.number_format = '#,##0.00'
                        
                        # Set deviation to 0.0 as requested
                        deviation_cell = template_sheet.cell(row=subtotal_row, column=11)
                        deviation_cell.value = 0.0
                        deviation_cell.font = Font(name='Calibri', size=11, bold=True)
                        deviation_cell.alignment = Alignment(horizontal='right', vertical='center')
                        deviation_cell.number_format = '#,##0.00'
                    
                    # Add light gray background to the entire subtotal row for better visibility
                    subtotal_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    for col in range(1, 12):
                        cell = template_sheet.cell(row=subtotal_row, column=col)
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        cell.border = subtotal_border
                    
                    row_count += 1  # Move to the next row
        
        # Determine the conclusion section start row based on audit approach and actual data
        conclusion_start_row = 71
        if audit_approach == "full":
            # For full audit approach, conclusion starts after the last data row
            # Add 5 rows of padding after the last data row
            conclusion_start_row = start_row + row_count + 5
            # Ensure it's at least row 150 for consistency
            conclusion_start_row = max(conclusion_start_row, 150)
        
        # Update conclusion section with account totals in a structured table format
        # First, ensure the header row spans all columns
        try:
            template_sheet.merge_cells(f'A{conclusion_start_row}:K{conclusion_start_row}')
        except:
            pass  # Ignore if cells are already merged
        
        # For each account, create a separate column in the conclusion table
        accounts = list(total_by_account.keys())
        
        # Define border style for consistency
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Define orange fill for the account cells
        orange_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
        
        # Add summary statistics for all operations
        total_operations = len(operations_df)
        total_amount = operations_df["Сума"].sum() if "Сума" in operations_df.columns else 0
        
        # Add a summary row before the conclusion section
        summary_row = conclusion_start_row - 3
        template_sheet.cell(row=summary_row, column=1).value = "ОБОБЩЕНА СТАТИСТИКА:"
        template_sheet.cell(row=summary_row, column=1).font = Font(name='Calibri', size=12, bold=True)
        template_sheet.cell(row=summary_row, column=1).alignment = Alignment(wrap_text=True, vertical='center')
        
        template_sheet.cell(row=summary_row+1, column=1).value = f"Общ брой операции: {total_operations}"
        template_sheet.cell(row=summary_row+1, column=1).font = Font(name='Calibri', size=11)
        
        template_sheet.cell(row=summary_row+1, column=4).value = f"Обща сума: {total_amount:.2f} лв."
        template_sheet.cell(row=summary_row+1, column=4).font = Font(name='Calibri', size=11)
        template_sheet.cell(row=summary_row+1, column=4).alignment = Alignment(horizontal='left')
        
        # Create a table-like structure for the conclusion that starts from column A
        # Process up to 4 accounts
        for i, account in enumerate(accounts[:4]):
            total = total_by_account[account]
            
            # Helper function to safely set cell value, handling merged cells
            def safe_set_cell_value(sheet, row, col, value, fill, alignment, border):
                cell = sheet.cell(row=row, column=col)
                
                # Check if this is a merged cell
                is_merged = False
                for merged_range in sheet.merged_cells.ranges:
                    # Get the range coordinates
                    min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
                    
                    # Check if our cell is within this range
                    if (min_row <= row <= max_row) and (min_col <= col <= max_col):
                        # If it's a merged cell, we need to use the top-left cell
                        if (row, col) != (min_row, min_col):
                            # This is not the top-left cell, so we need to use that one instead
                            cell = sheet.cell(row=min_row, column=min_col)
                        is_merged = True
                        break
                
                # Now set the value and styling
                cell.value = value
                cell.fill = fill
                cell.alignment = alignment
                cell.border = border
                
                return cell
            
            # Row 1: Checked documents total
            safe_set_cell_value(
                template_sheet,
                conclusion_start_row+1,
                i+1,
                f"Обща сума проверени документи по Кт на {account} - {total:.2f} лв.",
                orange_fill,
                Alignment(wrap_text=True, vertical='center'),
                border_style
            )
            
            # Row 2: Total sum by account
            safe_set_cell_value(
                template_sheet,
                conclusion_start_row+2,
                i+1,
                f"Обща сума по Кт на {account} - {total:.2f}лв.",
                orange_fill,
                Alignment(wrap_text=True, vertical='center'),
                border_style
            )
            
            # Row 3: Verification statement
            safe_set_cell_value(
                template_sheet,
                conclusion_start_row+3,
                i+1,
                "Стойността се равнява на тази по Об.ведомост и Гл.кн.",
                orange_fill,
                Alignment(wrap_text=True, vertical='center'),
                border_style
            )
        
        # Set the "НЕПРИЛОЖИМО" text in a merged cell for the error projection section
        try:
            template_sheet.merge_cells(f'A{conclusion_start_row+5}:K{conclusion_start_row+5}')
        except:
            pass
        
        safe_set_cell_value(
            template_sheet,
            conclusion_start_row+5,
            1,
            "НЕПРИЛОЖИМО",
            orange_fill,
            Alignment(wrap_text=True, vertical='center'),
            border_style
        )
        
        # Set the СНОН text in a merged cell
        try:
            template_sheet.merge_cells(f'A{conclusion_start_row+8}:K{conclusion_start_row+8}')
        except:
            pass
        
        safe_set_cell_value(
            template_sheet,
            conclusion_start_row+8,
            1,
            "Не са констатирани съществени неточности, отклонения и несъответствия при осчетоводяване на продажбите.",
            orange_fill,
            Alignment(wrap_text=True, vertical='center'),
            border_style
        )
        
        # Generate dynamic conclusion text based on the data and account type
        conclusion_text = ""
        
        # Get unique account types and the main account being analyzed
        account_types = set()
        main_account_being_analyzed = None
        
        # Determine which accounts to focus on based on account_type
        if account_type == "debit":
            # For debit reports, focus on debit accounts
            account_column = "Дт с/ка"
        else:
            # For credit reports or default, focus on credit accounts
            account_column = "Кт с/ка"
            
        # Extract the main account being analyzed from the data
        if account_column in operations_df.columns:
            accounts = operations_df[account_column].dropna().unique()
            if len(accounts) > 0:
                # Get the first account and extract its main part (first 3 digits)
                main_account = str(accounts[0])
                if '/' in main_account:
                    main_account = main_account.split('/')[0]
                if main_account and len(main_account) >= 3:
                    main_account_being_analyzed = main_account[:3]
        
        # Also collect all account types for reference
        for account_col in ["Дт с/ка", "Кт с/ка"]:
            if account_col in operations_df.columns:
                for account in operations_df[account_col].dropna().unique():
                    account_str = str(account)
                    if account_str and len(account_str) >= 3:
                        if '/' in account_str:
                            account_str = account_str.split('/')[0]
                        account_types.add(account_str[:3])
        
        # Generate conclusion based on the main account being analyzed
        if main_account_being_analyzed:
            # Reference the specific account being analyzed in the conclusion
            if main_account_being_analyzed in ['702', '703']:
                conclusion_text += f"При анализа на сметка {main_account_being_analyzed} се установи, че се извършват продажби на софтуерни услуги "
            elif main_account_being_analyzed == '704':
                conclusion_text += f"При анализа на сметка {main_account_being_analyzed} се установи, че се префактурират разходи за издръжка на предприятието "
            elif main_account_being_analyzed.startswith('4'):
                conclusion_text += f"При анализа на сметка {main_account_being_analyzed} се установи, че се отразяват разчетни взаимоотношения "
            elif main_account_being_analyzed.startswith('5'):
                conclusion_text += f"При анализа на сметка {main_account_being_analyzed} се установи, че се отразяват финансови взаимоотношения "
            elif main_account_being_analyzed.startswith('6'):
                conclusion_text += f"При анализа на сметка {main_account_being_analyzed} се установи, че се отразяват разходи "
            else:
                conclusion_text += f"При анализа на сметка {main_account_being_analyzed} се установи, че се извършват стандартни счетоводни операции "
        else:
            # Fallback to the old logic if we can't determine the main account
            if '702' in account_types or '703' in account_types:
                conclusion_text += "Извършват се продажби на софтуерни услуги "
            
            if '704' in account_types:
                if conclusion_text:
                    conclusion_text += "и се префактурират "
                else:
                    conclusion_text += "Префактурират се "
                conclusion_text += "разходи за издръжка на предприятието "
            
            # Add general conclusion if nothing specific was added
            if not conclusion_text:
                conclusion_text = "Извършват се стандартни счетоводни операции. "
        
        # Add general conclusion about accounting practices
        if conclusion_text and not conclusion_text.endswith(". "):
            conclusion_text += "за сметка на фирмата-майка. "
            
        conclusion_text += "Използват се подходящи счетоводни сметки с детайлна аналитичност. Записванията се отразяват своевременно в регистрите на дружеството."
        
        # Set the conclusion text in a merged cell
        try:
            template_sheet.merge_cells(f'A{conclusion_start_row+13}:K{conclusion_start_row+13}')
        except:
            pass
        
        safe_set_cell_value(
            template_sheet,
            conclusion_start_row+13,
            1,
            conclusion_text,
            orange_fill,
            Alignment(wrap_text=True, vertical='center'),
            border_style
        )
        
        # Create a BytesIO object to store the result
        result = io.BytesIO()
        
        # Save the workbook to the BytesIO object
        template_wb.save(result)
        
        # Reset the file position to the beginning
        result.seek(0)
        
        return result
    
    def wrap_and_upload_excel(self,
                             s3_key: str,
                             company_name: str = "Форт България ЕООД",
                             year: str = None,
                             audit_approach: str = "statistical",
                             account_type: str = None) -> Optional[str]:
        """
        Download an Excel file from S3, wrap it with a template, and upload it back to S3
        
        Args:
            s3_key: S3 key of the Excel file to wrap
            company_name: Name of the company to include in the template
            year: Year to include in the template
            audit_approach: The audit approach to use (default: "statistical")
            account_type: The type of account being analyzed ("debit" or "credit")
            
        Returns:
            S3 key of the wrapped Excel file if successful, None otherwise
        """
        try:
            # Download the Excel file from S3
            excel_content = self.s3_service.download_file(s3_key)
            if not excel_content:
                print(f"Error downloading file from S3: {s3_key}")
                return None
            
            # Wrap the Excel file with the template
            wrapped_excel = self.wrap_excel_with_template(
                excel_content,
                company_name=company_name,
                year=year,
                audit_approach=audit_approach,
                account_type=account_type
            )
            
            # Generate a new S3 key for the wrapped file
            # Keep the same directory structure but add "_wrapped" to the filename
            path_parts = s3_key.split('/')
            filename = path_parts[-1]
            filename_parts = filename.split('.')
            wrapped_filename = f"{filename_parts[0]}_wrapped.{filename_parts[1]}"
            path_parts[-1] = wrapped_filename
            wrapped_s3_key = '/'.join(path_parts)
            
            # Upload the wrapped Excel file to S3
            success, message = self.s3_service.upload_file(wrapped_excel, wrapped_s3_key)
            
            if success:
                return wrapped_s3_key
            else:
                print(f"Error uploading wrapped file to S3: {message}")
                return None
                
        except Exception as e:
            print(f"Error wrapping Excel file: {str(e)}")
            import traceback
            traceback.print_exc()
            return None