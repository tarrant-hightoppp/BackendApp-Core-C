from datetime import datetime
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from app.services.excel_report.cell_utils import CellUtils


class TemplateGenerator:
    """Class for generating Excel templates for accounting reports"""

    @staticmethod
    def create_template_workbook(
        company_name: str = "Форт България ЕООД",
        year: Optional[str] = None,
        auditor_name: str = "ПРИМА ФИНАНС КОНСУЛТИНГ ЕООД",
        audit_approach: str = "statistical"
    ) -> Workbook:
        """
        Create a new workbook with the C700 template structure
        
        Args:
            company_name: Name of the company to include in the template
            year: Year to include in the template (default: current year)
            auditor_name: Name of the auditor company
            audit_approach: The audit approach to use (default: "statistical")
            
        Returns:
            Workbook object with the template structure
        """
        # If year is not provided, use current year
        if year is None:
            year = str(datetime.now().year)
            
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Define styles
        header_font = Font(name='Calibri', size=11, bold=True)
        normal_font = Font(name='Calibri', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        # Define a thicker border for the C-D column separation
        thick_right_border = Border(
            left=Side(style='thin'),
            right=Side(style='thick'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        # Colors based on the C700.xlsx template - exact color codes
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        teal_fill = PatternFill(start_color="33CCCC", end_color="33CCCC", fill_type="solid")  # Exact teal color for header rows
        light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        footer_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White for footer
        
        # Set column widths - match the exact widths from the template
        column_widths = {
            'A': 15.43,   # № по ред
            'B': 22.14,  # Документ №/Дата
            'C': 10.0,   # Рег. №
            'D': 10.0,   # Дт с/ка
            'E': 25.71,  # Аналитична сметка (Дт)
            'F': 10.0,   # Кт с/ка
            'G': 25.71,  # Аналитична сметка (Кт)
            'H': 15.0,   # Сума
            'I': 30.0,   # Обяснение/Обоснование
            'J': 15.0,   # Установена сума при одита
            'K': 15.0,   # Отклонение
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
            
        # Set row heights - match the exact heights from the template
        row_heights = {
            # Header section
            1: 38, 2: 20.25, 3: 50.0, 4: 40, 5: 40, 6: 40, 7: 30.0, 8: 20.25,
            # Audit purpose section
            9: 20.25, 10: 30.0, 11: 20.25, 12: 20.25, 13: 30.0,
            # Approach section
            15: 20.25, 16: 20.25, 17: 20.25, 18: 20.25,
            # Operations header
            26: 30.0,
            #Footer \\^.^//
            72: 30, 74: 30, 79: 30,84: 35,
        }
        
        # Set specific row heights
        for row, height in row_heights.items():
            ws.row_dimensions[row].height = height
            
        # Set default height for other rows
        for i in range(1, 100):
            if i not in row_heights:
                if i >= 71:  # Footer section
                    ws.row_dimensions[i].height = 18.0  # Taller rows for footer
                else:
                    ws.row_dimensions[i].height = 15.0
        
        # Create merged cells for the header section based on the exact structure in C700.xlsx
        # First row
        ws.merge_cells('A1:A1')
        ws.merge_cells('B1:C1')
        ws.merge_cells('D1:E1')
        ws.merge_cells('F1:G1')
        
        # Second row
        ws.merge_cells('A2:A2')
        ws.merge_cells('B2:C2')
        ws.merge_cells('D2:E2')
        ws.merge_cells('F2:G2')
        
        # Third row
        ws.merge_cells('A3:A3')
        ws.merge_cells('B3:C3')
        ws.merge_cells('D3:E3')
        ws.merge_cells('F3:G3')
        
        # Fourth row
        ws.merge_cells('A4:A4')
        ws.merge_cells('B4:C4')
        ws.merge_cells('D4:E4')
        ws.merge_cells('F4:G4')
        
        # Fifth row
        ws.merge_cells('A5:A5')
        ws.merge_cells('B5:C5')
        ws.merge_cells('D5:E5')
        ws.merge_cells('F5:G5')
        
        # Sixth row
        ws.merge_cells('A6:A6')
        ws.merge_cells('B6:C6')
        ws.merge_cells('D6:E6')
        ws.merge_cells('F6:G6')
        
        # Seventh row
        ws.merge_cells('A7:C7')
        ws.merge_cells('D7:G7')
        
        # Eighth row
        ws.merge_cells('A8:C8')
        ws.merge_cells('D8:G8')
        
        # Header section - First part (rows 1-8)
        # Row 1
        ws['A1'] = "ОДИТЪТ СЕ ИЗВЪРШВА ОТ"
        ws['B1'] = auditor_name
        ws['D1'] = "ДОКУМЕНТ"
        ws['F1'] = "С 700  Тест  по  същество  на   приходи"
        
        # Row 2
        ws['A2'] = "АДРЕС"
        ws['B2'] = "СОФИЯ, УЛ. ЦАР ШИШМАН 17"
        ws['D2'] = "ОДИТИРАНО ПРЕДПРИЯТИЕ"
        ws['F2'] = company_name
        
        # Row 3
        ws['A3'] = "БУЛСТАТ"
        ws['B3'] = "121100122"
        ws['D3'] = "АДРЕС"
        ws['F3'] = "гр.София, жк Младост, Бизнес\nЦентър \"Капитал Форт\", бул. \"Цариградско шосе\" No 90, ет. 13"
        
        # Row 4
        ws['A4'] = "РЪКОВОДИТЕЛ ЕКИП"
        ws['B4'] = "ВАСИЛ КРЪСТЕВ КАЛАЙДЖИЕВ"
        ws['D4'] = "БУЛСТАТ"
        ws['F4'] = "203576042"
        
        # Row 5
        ws['A5'] = "ДИПЛОМА НОМЕР"
        ws['B5'] = "409"
        ws['D5'] = "ПРОВЕРЯВАН ПЕРИОД"
        ws['F5'] = year
        
        # Row 6
        ws['A6'] = "ДАТА НА ИЗГОТВЯНЕ"
        ws['B6'] = datetime.now().strftime("%d/%m/%Y")
        ws['D6'] = "ДАТА НА ПРОВЕРКА"
        ws['F6'] = datetime.now().strftime("%d/%m/%Y")
        
        # Row 7
        ws['A7'] = "ИЗГОТВИЛ ИЛИ НАНЕСЪЛ ПОПРАВКИ (ОДИТОР/ПОМОЩНИК ОДИТОР,АСИСТЕНТ)"
        ws['D7'] = "ПРОВЕРИЛ (ОТГОВОРЕН ОДИТОР)"
        
        # Row 8
        ws['A8'] = "ИС"
        ws['D8'] = "ВК"
        
        # Apply styles to header section - teal background for header rows
        for row in range(1, 9):
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.font = normal_font
                cell.alignment = Alignment(wrap_text=True, vertical='center')
                # Apply thicker border between columns C and D in the header
                if col == 3:  # Column C
                    cell.border = thick_right_border
                else:
                    cell.border = border
                cell.fill = teal_fill  # Apply teal background to header rows
        
        # Audit purpose section - merge cells for all columns A-G
        for row in range(9, 14):
            ws.merge_cells(f'A{row}:G{row}')
        
        # Audit purpose section - content
        ws['A9'] = "Цел   на одиторската  процедура "
        ws['A10'] = "Целта  на   одиторската  процедура  е да  установи   СНОН  при  признаването ,  оценката , последваща  оценка ,  класификация  и представяне  "
        ws['A11'] = "на приходи  , включително  и  финансови  ."
        ws['A12'] = "Приложени одиторсик процедури :"
        ws['A13'] = "факт.проверка  на договори и др.документация;повторно изчисление на салдо ;равнение на сметката с ФО ;проучващи запитвания"
        
        # Apply styles to audit purpose section
        for row in range(9, 14):
            cell = ws.cell(row=row, column=1)
            cell.font = normal_font
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')  # Center horizontally
            cell.border = border
        
        # Approach section - merge cells
        # Row 15 - header
        ws.merge_cells('C15:G15')
        
        # Rows 16-18 - options
        for row in range(16, 19):
            ws.merge_cells(f'C{row}:G{row}')
        
        # Approach section - content
        ws['C15'] = "  Избран подход за тест по  същество "
        ws['C16'] = "проверка на 100 %  на  популация "
        ws['C17'] = "проверка   на  избрани  обекти  на  популация  "
        # Set the X mark based on the audit approach
        if audit_approach == "full":
            ws['A16'] = "X"  # 100% population check
            ws['A17'] = ""
            ws['A18'] = ""
        elif audit_approach == "selected":
            ws['A16'] = ""
            ws['A17'] = "X"  # Check of selected population objects
            ws['A18'] = ""
        else:  # statistical (default)
            ws['A16'] = ""
            ws['A17'] = ""
            ws['A18'] = "X"  # Statistical audit sampling
        ws['C18'] = "одиторска  извадка   - статистическа "
        
        # Apply styles to approach section
        # Style for the header
        cell = ws.cell(row=15, column=3)
        cell.font = normal_font
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.border = border
        
        # Style for the checkboxes
        for row in [16, 17, 18]:
            cell = ws.cell(row=row, column=1)
            cell.font = normal_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Style for the options
        for row in range(16, 19):
            cell = ws.cell(row=row, column=3)
            cell.font = normal_font
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.border = border
        
        # Table header for operations - row 26
        headers = [
            "Документ №/Дата", "Рег. №", "Дт с/ка", "Аналитична сметка",
            "Кт с/ка", "Аналитична сметка", "Сума", "Обяснителен текст",
            "Установена сума   при одита ", "Отклонение  "
        ]
        
        # Add column headers for operations table (row 26)
        for i, header in enumerate(["№ по ред"] + headers):
            cell = ws.cell(row=26, column=i+1)
            cell.value = header
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            cell.fill = teal_fill  # Use teal color for the operations header row
            
        # Add empty rows for operations
        # For "full" audit approach, we need more rows - allocate 500 rows to be safe
        max_row = 500 if audit_approach == "full" else 71
        for row in range(27, max_row):
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.border = border
        
        # Add conclusion section headers
        # For "full" audit approach, the conclusion section starts after the operations
        conclusion_start_row = max_row
        
        # Create the conclusion table header
        ws.merge_cells(f'A{conclusion_start_row}:K{conclusion_start_row}')
        
        # Merge A-B for the label columns
        for i in range(1, 15):
            row = conclusion_start_row + i
            ws.merge_cells(f'A{row}:B{row}')
        
        # Conclusion section content - headers only (values will be populated dynamically)
        ws[f'A{conclusion_start_row}'] = "ЗАКЛЮЧЕНИЯ :"
        ws[f'A{conclusion_start_row+1}'] = "Обща  сума/брой   на  проверени   документи "
        ws[f'A{conclusion_start_row+2}'] = "Обща  сума / брой на обекти в  популация "
        ws[f'A{conclusion_start_row+3}'] = "Равнение на  ст/ст на популация с оборотна ведомости; глава  книга ; ФО "
        ws[f'A{conclusion_start_row+4}'] = "Проектиране на грешката "
        ws[f'A{conclusion_start_row+5}'] = "Проектиране на грешката "
        ws[f'C{conclusion_start_row+5}'] = "НЕПРИЛОЖИМО"
        ws[f'A{conclusion_start_row+6}'] = ""
        ws[f'A{conclusion_start_row+7}'] = ""
        ws[f'A{conclusion_start_row+8}'] = "Констатирани  СНОН "
        ws[f'C{conclusion_start_row+8}'] = "Не са констатирани съществени неточности, отклонения и несъответствия при осчетоводяване на продажбите."
        ws[f'A{conclusion_start_row+13}'] = "Други  заключения "
        
        # Apply styles to conclusion section - with proper formatting
        # Header row (ЗАКЛЮЧЕНИЯ) with improved formatting
        for col in range(1, 12):
            cell = ws.cell(row=conclusion_start_row, column=col)
            cell.font = Font(name='Calibri', size=12, bold=True)  # Slightly larger font
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')  # Center align
            cell.border = border
            cell.fill = teal_fill  # Use teal color for the conclusion header
        
        # Content rows with improved formatting
        for i in range(1, 20):  # Format 20 rows after conclusion header
            row = conclusion_start_row + i
            
            # Format label column (A-B merged)
            label_cell = ws.cell(row=row, column=1)
            if not label_cell.font:
                label_cell.font = normal_font
            label_cell.alignment = Alignment(wrap_text=True, vertical='center')
            label_cell.border = border
            
            # Apply different formatting to different sections
            if i in [1, 2, 3]:  # First section - Account totals
                label_cell.fill = light_gray_fill
                label_cell.font = Font(name='Calibri', size=11, bold=True)  # Make labels bold
            elif i in [4, 5, 6, 7]:  # Проектиране на грешката section
                label_cell.fill = light_gray_fill
                if i == 4:  # Section header
                    label_cell.font = Font(name='Calibri', size=11, bold=True)
            elif i in [8, 9, 10, 11, 12]:  # Констатирани СНОН section
                label_cell.fill = light_gray_fill
                if i == 8:  # Section header
                    label_cell.font = Font(name='Calibri', size=11, bold=True)
            elif i == 13:  # Други заключения section
                label_cell.fill = light_gray_fill
                label_cell.font = Font(name='Calibri', size=11, bold=True)
            
            # Format content cells (C-K) - individual cells for each account
            for col in range(3, 12):
                cell = ws.cell(row=row, column=col)
                if not cell.font:
                    cell.font = normal_font
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                cell.border = border
                cell.fill = footer_fill
        
        return wb