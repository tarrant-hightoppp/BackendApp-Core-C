from typing import Dict, List, Optional
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from app.services.excel_report.cell_utils import CellUtils


class ConclusionGenerator:
    """Class for generating conclusion sections in Excel reports"""

    @staticmethod
    def generate_conclusion_text(
        main_account_being_analyzed: Optional[str] = None,
        account_types: Optional[set] = None
    ) -> str:
        """
        Generate conclusion text based on the main account being analyzed
        
        Args:
            main_account_being_analyzed: The main account being analyzed (first 3 digits)
            account_types: Set of all account types found in the data
            
        Returns:
            Conclusion text
        """
        conclusion_text = ""
        
        # Generate conclusion based on the main account being analyzed
        if main_account_being_analyzed:
            # Reference the specific account being analyzed in the conclusion
            if main_account_being_analyzed in ['702', '703']:
                conclusion_text += f"При анализа на сметка {main_account_being_analyzed} се установи, че се извършват продажби на софтуерни услуги "
            elif main_account_being_analyzed == '704':
                conclusion_text += f"При анализа на сметка {main_account_being_analyzed} се установи, че се префактурират разходи за издръжка на предприятието "
            elif main_account_being_analyzed.startswith('4'):
                conclusion_text += f"При анализа на сметка {main_account_being_analyzed} се установи, че се отразяват разчетни взаимоотношения "
            elif main_account_being_analyzed.startswith('5'):
                conclusion_text += f"При анализа на сметка {main_account_being_analyzed} се установи, че се отразяват финансови взаимоотношения "
            elif main_account_being_analyzed.startswith('6'):
                conclusion_text += f"При анализа на сметка {main_account_being_analyzed} се установи, че се отразяват разходи "
            else:
                conclusion_text += f"При анализа на сметка {main_account_being_analyzed} се установи, че се извършват стандартни счетоводни операции "
        else:
            # Fallback to the old logic if we can't determine the main account
            if account_types and ('702' in account_types or '703' in account_types):
                conclusion_text += "Извършват се продажби на софтуерни услуги "
            
            if account_types and '704' in account_types:
                if conclusion_text:
                    conclusion_text += "и се префактурират "
                else:
                    conclusion_text += "Префактурират се "
                conclusion_text += "разходи за издръжка на предприятието "
            
            # Add general conclusion if nothing specific was added
            if not conclusion_text:
                conclusion_text = "Извършват се стандартни счетоводни операции. "
        
        # Add general conclusion about accounting practices
        if conclusion_text and not conclusion_text.endswith(". "):
            conclusion_text += "за сметка на фирмата-майка. "
            
        conclusion_text += "Използват се подходящи счетоводни сметки с детайлна аналитичност. Записванията се отразяват своевременно в регистрите на дружеството."
        
        return conclusion_text

    @staticmethod
    def populate_conclusion_section(
        template_sheet: Worksheet,
        conclusion_start_row: int,
        total_by_account: Dict[str, float],
        total_operations: int,
        total_amount: float
    ):
        """
        Populate the conclusion section of the template with account totals
        
        Args:
            template_sheet: The worksheet to modify
            conclusion_start_row: The starting row for the conclusion section
            total_by_account: Dictionary with account numbers as keys and total amounts as values
            total_operations: Total number of operations
            total_amount: Total amount across all operations
        """
        # Define border style for consistency
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Define orange fill for the account cells
        orange_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
        
        # Add a summary row before the conclusion section
        summary_row = conclusion_start_row - 3
        
        # Merge cells A-M for the summary header
        try:
            template_sheet.merge_cells(f'A{summary_row}:M{summary_row}')
        except:
            pass
            
        CellUtils.safe_set_cell_value(
            template_sheet,
            summary_row,
            1,
            "ОБОБЩЕНА СТАТИСТИКА:",
            font=Font(name='Calibri', size=48, bold=True),
            alignment=Alignment(wrap_text=True, vertical='center', horizontal='center')
        )
        
        # Merge cells A-M for the summary content row
        try:
            template_sheet.merge_cells(f'A{summary_row+1}:M{summary_row+1}')
        except:
            pass
            
        CellUtils.safe_set_cell_value(
            template_sheet,
            summary_row+1,
            1,
            f"Общ брой операции: {total_operations}     |     Обща сума: {total_amount:.2f} лв.",
            font=Font(name='Calibri', size=14, bold=True),
            alignment=Alignment(horizontal='center', vertical='center')
        )
        
        # Merge cells A-M for the first few rows and populate with account summaries
        accounts = list(total_by_account.keys())
        
        # Row 1: Checked documents total
        try:
            template_sheet.merge_cells(f'A{conclusion_start_row+1}:M{conclusion_start_row+1}')
        except:
            pass
        
        account_summary_text = ""
        for i, account in enumerate(accounts[:4]):
            total = total_by_account[account]
            if i > 0:
                account_summary_text += "; "
            account_summary_text += f"Обща сума проверени документи по Кт на {account} - {total:.2f} лв."
        
        CellUtils.safe_set_cell_value(
            template_sheet,
            conclusion_start_row+1,
            1,
            account_summary_text,
            orange_fill,
            Alignment(wrap_text=True, vertical='center', horizontal='center'),
            border_style
        )
        
        # Row 2: Total sum by account
        try:
            template_sheet.merge_cells(f'A{conclusion_start_row+2}:M{conclusion_start_row+2}')
        except:
            pass
            
        account_totals_text = ""
        for i, account in enumerate(accounts[:4]):
            total = total_by_account[account]
            if i > 0:
                account_totals_text += "; "
            account_totals_text += f"Обща сума по Кт на {account} - {total:.2f}лв."
        
        CellUtils.safe_set_cell_value(
            template_sheet,
            conclusion_start_row+2,
            1,
            account_totals_text,
            orange_fill,
            Alignment(wrap_text=True, vertical='center', horizontal='center'),
            border_style
        )
        
        # Row 3: Verification statement
        try:
            template_sheet.merge_cells(f'A{conclusion_start_row+3}:M{conclusion_start_row+3}')
        except:
            pass
            
        CellUtils.safe_set_cell_value(
            template_sheet,
            conclusion_start_row+3,
            1,
            "Стойността се равнява на тази по Об.ведомост и Гл.кн.",
            orange_fill,
            Alignment(wrap_text=True, vertical='center', horizontal='center'),
            border_style
        )
        
        # Merge cells for rows 4, 6, 7, 9-12 (empty or placeholder rows)
        for row_offset in [4, 6, 7, 9, 10, 11, 12]:
            try:
                template_sheet.merge_cells(f'A{conclusion_start_row+row_offset}:M{conclusion_start_row+row_offset}')
            except:
                pass
        
        # Set the "НЕПРИЛОЖИМО" text in a merged cell for the error projection section
        try:
            template_sheet.merge_cells(f'A{conclusion_start_row+5}:M{conclusion_start_row+5}')
        except:
            pass
        
        CellUtils.safe_set_cell_value(
            template_sheet,
            conclusion_start_row+5,
            1,
            "НЕПРИЛОЖИМО",
            orange_fill,
            Alignment(wrap_text=True, vertical='center', horizontal='center'),
            border_style
        )
        
        # Set the СНОН text in a merged cell
        try:
            template_sheet.merge_cells(f'A{conclusion_start_row+8}:M{conclusion_start_row+8}')
        except:
            pass
        
        CellUtils.safe_set_cell_value(
            template_sheet,
            conclusion_start_row+8,
            1,
            "Не са констатирани съществени неточности, отклонения и несъответствия при осчетоводяване на продажбите.",
            orange_fill,
            Alignment(wrap_text=True, vertical='center', horizontal='center'),
            border_style
        )