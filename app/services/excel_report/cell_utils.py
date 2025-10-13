from typing import Optional
from openpyxl.styles import Alignment, Font, Border, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


class CellUtils:
    """Utility class for handling Excel cell operations, especially for merged cells"""

    @staticmethod
    def safe_set_cell_value(
        sheet: Worksheet, 
        row: int, 
        col: int, 
        value: any, 
        fill: Optional[PatternFill] = None, 
        alignment: Optional[Alignment] = None, 
        border: Optional[Border] = None, 
        number_format: Optional[str] = None, 
        font: Optional[Font] = None
    ):
        """
        Safely set a cell value, handling merged cells by finding the top-left cell
        
        Args:
            sheet: The worksheet to modify
            row: Row number
            col: Column number
            value: Value to set
            fill: Cell fill pattern (optional)
            alignment: Cell alignment (optional)
            border: Cell border (optional)
            number_format: Cell number format (optional)
            font: Cell font (optional)
            
        Returns:
            The cell that was modified
        """
        # Check if this is a merged cell
        is_merged = False
        cell = sheet.cell(row=row, column=col)
        
        for merged_range in sheet.merged_cells.ranges:
            # Get the range coordinates
            min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
            
            # Check if our cell is within this range
            if (min_row <= row <= max_row) and (min_col <= col <= max_col):
                # If it's a merged cell, we need to use the top-left cell
                if (row, col) != (min_row, min_col):
                    # This is not the top-left cell, so we need to use that one instead
                    cell = sheet.cell(row=min_row, column=min_col)
                is_merged = True
                break
        
        # Now set the value and styling
        cell.value = value
        
        if fill is not None:
            cell.fill = fill
        if alignment is not None:
            cell.alignment = alignment
        if border is not None:
            cell.border = border
        if number_format is not None:
            cell.number_format = number_format
        if font is not None:
            cell.font = font
        
        return cell