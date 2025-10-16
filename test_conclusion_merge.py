import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from app.services.excel_report.template_wrapper import ExcelTemplateWrapper

# Create test data
test_data = {
    "№ по ред": [1, 2, 3, 4, 5],
    "Вид документ": ["ФИ", "ФИ", "ФИ", "ФИ", "ФИ"],
    "Документ №": ["0000000001", "0000000002", "0000000003", "0000000004", "0000000005"],
    "Дата": ["01.01.2023", "02.01.2023", "03.01.2023", "04.01.2023", "05.01.2023"],
    "Дт с/ка": ["401", "401", "401", "401", "401"],
    "Аналитична сметка/Партньор (Дт)": ["Test Partner 1", "Test Partner 2", "Test Partner 3", "Test Partner 4", "Test Partner 5"],
    "Кт с/ка": ["702", "702", "703", "703", "704"],
    "Аналитична сметка/Партньор (Кт)": ["Service 1", "Service 2", "Service 3", "Service 4", "Service 5"],
    "Сума": [1000.00, 2000.00, 1500.00, 2500.00, 3000.00],
    "Обяснение/Обоснование": ["Test explanation 1", "Test explanation 2", "Test explanation 3", "Test explanation 4", "Test explanation 5"],
    "Установена сума при одита": [1000.00, 2000.00, 1500.00, 2500.00, 3000.00],
    "Отклонение": [0.00, 0.00, 0.00, 0.00, 0.00],
    "Установено контролно действие при одита": ["OK", "OK", "OK", "OK", "OK"]
}

# Create DataFrame
df = pd.DataFrame(test_data)

# Save to Excel
test_file = "test_operations.xlsx"
df.to_excel(test_file, index=False)

# Create wrapper instance
wrapper = ExcelTemplateWrapper()

# Read the test file
with open(test_file, 'rb') as f:
    excel_content = f.read()

# Wrap with template
wrapped_content = wrapper.wrap_excel_with_template(
    excel_content,
    company_name="Test Company",
    year="2023",
    audit_approach="full",
    account_type="credit"
)

# Save wrapped file
output_file = "test_wrapped_output.xlsx"
with open(output_file, 'wb') as f:
    f.write(wrapped_content.getvalue())

print(f"Wrapped file saved as: {output_file}")

# Load and check the wrapped file
wb = load_workbook(output_file)
ws = wb.active

# Find the conclusion section
conclusion_row = None
for row in range(1, ws.max_row + 1):
    cell_value = ws.cell(row=row, column=1).value
    if cell_value and "ЗАКЛЮЧЕНИЯ" in str(cell_value):
        conclusion_row = row
        break

if conclusion_row:
    print(f"\nConclusion section found at row: {conclusion_row}")
    
    # Check merged cells in the conclusion section
    print("\nChecking merged cells in conclusion section:")
    for i in range(1, 15):
        row = conclusion_row + i
        # Check if cells A-M are merged
        is_merged = False
        for merged_range in ws.merged_cells.ranges:
            if f"A{row}" in merged_range:
                print(f"Row {row}: Merged range {merged_range}")
                is_merged = True
                break
        if not is_merged:
            print(f"Row {row}: NOT MERGED")
else:
    print("ERROR: Conclusion section not found!")

# Clean up test files
import os
os.remove(test_file)
print(f"\nTest completed. Check {output_file} to verify the merged cells.")