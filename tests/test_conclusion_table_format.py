import io
import os
import pytest
from openpyxl import load_workbook
from app.services.excel_template_wrapper import ExcelTemplateWrapper

@pytest.fixture
def sample_excel_data():
    """Create a sample Excel file with operations data for testing"""
    import pandas as pd
    
    # Create a DataFrame with sample operations
    data = {
        "№ по ред": [1, 2, 3, 4],
        "Документ №": ["DOC001", "DOC002", "DOC003", "DOC004"],
        "Дата": ["01.01.2025", "02.01.2025", "03.01.2025", "04.01.2025"],
        "Дт с/ка": ["411", "411", "411", "411"],
        "Аналитична сметка/Партньор (Дт)": ["Partner1", "Partner2", "Partner3", "Partner4"],
        "Кт с/ка": ["702001", "702002", "705001", "709101"],
        "Аналитична сметка/Партньор (Кт)": ["AnalyticalKt1", "AnalyticalKt2", "AnalyticalKt3", "AnalyticalKt4"],
        "Сума": [601973.41, 367739.61, 500.00, 729.77],
        "Обяснение/Обоснование": ["Explanation1", "Explanation2", "Explanation3", "Explanation4"]
    }
    
    df = pd.DataFrame(data)
    
    # Save to a BytesIO object
    excel_data = io.BytesIO()
    df.to_excel(excel_data, index=False)
    excel_data.seek(0)
    
    return excel_data

def test_conclusion_table_format(sample_excel_data):
    """Test that the conclusion table has the correct format"""
    # Create an instance of ExcelTemplateWrapper
    wrapper = ExcelTemplateWrapper()
    
    # Wrap the sample Excel data with the template
    wrapped_excel = wrapper.wrap_excel_with_template(
        sample_excel_data,
        company_name="Test Company",
        year="2025",
        audit_approach="full",
        account_type="debit"
    )
    
    # Load the wrapped Excel file
    wb = load_workbook(wrapped_excel)
    ws = wb.active
    
    # Find the conclusion section
    conclusion_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == "ЗАКЛЮЧЕНИЯ :":
            conclusion_row = row
            break
    
    assert conclusion_row is not None, "Conclusion section not found"
    
    # Check that the conclusion header spans all columns
    merged_ranges = [str(merged_cell) for merged_cell in ws.merged_cells.ranges]
    assert f"A{conclusion_row}:K{conclusion_row}" in merged_ranges, "Conclusion header should span columns A-K"
    
    # Check that the first row after the header has the correct format
    # Check that we have separate cells for each account in the first row
    for col in range(1, 5):  # Columns A through D
        cell = ws.cell(row=conclusion_row+1, column=col)
        assert cell.value is not None, f"Cell {chr(64+col)}{conclusion_row+1} should have a value"
        assert "Обща сума проверени документи по Кт на" in str(cell.value), f"Cell {chr(64+col)}{conclusion_row+1} should contain account total"
    
    # Check that the НЕПРИЛОЖИМО section is merged
    assert any(f"A{conclusion_row+5}" in merged_range for merged_range in merged_ranges), "НЕПРИЛОЖИМО section should be merged"
    
    # Check that the СНОН section is merged
    assert any(f"A{conclusion_row+8}" in merged_range for merged_range in merged_ranges), "СНОН section should be merged"
    
    # Check that the conclusion text is in a merged cell
    assert any(f"A{conclusion_row+13}" in merged_range for merged_range in merged_ranges), "Conclusion text should be in a merged cell"
    
    # Check that the cells have the correct fill color
    orange_fill_hex = "FFCC99"
    for row in [conclusion_row+1, conclusion_row+2, conclusion_row+3]:
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            assert cell.fill.start_color.rgb == orange_fill_hex, f"Cell {chr(64+col)}{row} should have orange fill"
    
    # Check that summary statistics are present
    summary_row = conclusion_row - 3
    assert "ОБОБЩЕНА СТАТИСТИКА" in str(ws.cell(row=summary_row, column=1).value), "Summary statistics header should be present"
    assert "Общ брой операции" in str(ws.cell(row=summary_row+1, column=1).value), "Operation count should be present"
    assert "Обща сума" in str(ws.cell(row=summary_row+1, column=4).value), "Total amount should be present"
    
    print("Conclusion table format test passed!")

if __name__ == "__main__":
    # Run the test directly
    sample_data = sample_excel_data()
    test_conclusion_table_format(sample_data)
    print("All tests passed!")